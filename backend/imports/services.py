"""Vehicle Excel import validation and commit services."""

from __future__ import annotations

import itertools
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any

from django.core.exceptions import ValidationError as DjangoValidationError
from django.core.files.storage import default_storage
from django.db import IntegrityError, transaction
from django.utils import timezone
from django.utils.translation import gettext as _
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from audit.models import AuditLog
from imports.models import ImportJob
from mediafiles.models import MediaType
from mediafiles.services import create_media_file_from_upload
from vehicles.models import Vehicle, VehicleCategory, VehicleStatus


EXPECTED_COLUMNS = [
    "internal_number",
    "category",
    "manufacturer",
    "model",
    "serial_number",
    "license_plate",
    "current_odometer_km",
    "current_operating_hours",
    "current_location",
    "supplier",
    "notes",
]
# Only the data we cannot derive is required. ``internal_number`` is optional:
# when the column is missing or a cell is blank the fleet number is generated
# automatically (e.g. FZ-00001). ``category`` is optional too: unknown or blank
# categories fall back to the catch-all category below. Any expected column
# missing from the file is simply treated as blank for every row.
REQUIRED_COLUMNS = ["manufacturer", "model"]
SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}

# Catch-all category used when an imported row references an unknown/inactive
# category or leaves it blank. Created on demand at commit time.
FALLBACK_CATEGORY_NAME = "Sonstiges"

# Mirror the Vehicle model field limits so over-long values surface as readable
# row errors during validation instead of failing the model's full_clean() at
# commit time.
_MAX_FIELD_LENGTHS = {
    "internal_number": 80,
    "manufacturer": 120,
    "model": 120,
    "serial_number": 120,
    "license_plate": 40,
    "current_location": 255,
}

# Accept common German and English header spellings so files do not have to use
# the exact internal column keys. Keys are the canonical columns; values are the
# normalized header variants that map onto them.
COLUMN_ALIASES: dict[str, set[str]] = {
    "internal_number": {
        "internal_number",
        "interne_nummer",
        "fahrzeugnummer",
        "fahrzeug_nummer",
        "fahrzeug_nr",
        "fzg_nr",
        "fzg_nummer",
        "nummer",
    },
    "category": {
        "category",
        "kategorie",
        "fahrzeugkategorie",
        "fahrzeugart",
        "fahrzeugtyp",
        "art",
    },
    "manufacturer": {"manufacturer", "hersteller", "marke"},
    "model": {"model", "modell", "bezeichnung", "typ"},
    "serial_number": {
        "serial_number",
        "seriennummer",
        "serien_nr",
        "fahrgestellnummer",
        "fin",
    },
    "license_plate": {
        "license_plate",
        "kennzeichen",
        "kfz_kennzeichen",
        "amtliches_kennzeichen",
        "nummernschild",
    },
    "current_odometer_km": {
        "current_odometer_km",
        "kilometerstand",
        "km_stand",
        "kilometer",
        "km",
    },
    "current_operating_hours": {
        "current_operating_hours",
        "betriebsstunden",
        "betriebsstd",
        "stunden",
    },
    "current_location": {
        "current_location",
        "standort",
        "lagerort",
        "ort",
    },
    "supplier": {"supplier", "lieferant", "zulieferer"},
    "notes": {
        "notes",
        "notizen",
        "bemerkung",
        "bemerkungen",
        "anmerkung",
        "anmerkungen",
        "kommentar",
    },
}

# Reverse lookup: normalized header variant -> canonical column key.
_ALIAS_TO_COLUMN = {alias: column for column, aliases in COLUMN_ALIASES.items() for alias in aliases}


@dataclass(frozen=True)
class ImportValidationResult:
    row_count: int
    error_count: int
    result: dict[str, Any]


@transaction.atomic
def create_vehicle_import_job(*, uploaded_file, actor, request_meta: dict[str, str]) -> ImportJob:
    """Create media metadata, validate the workbook, and persist an ImportJob."""
    source_media = create_media_file_from_upload(
        uploaded_file=uploaded_file,
        actor=actor,
        media_type=MediaType.IMPORT,
        related_type="vehicle_import",
    )
    job = ImportJob.objects.create(
        import_type=ImportJob.ImportType.VEHICLES,
        source_media=source_media,
        created_by=actor,
    )

    validation = validate_vehicle_workbook(uploaded_file)
    job.row_count = validation.row_count
    job.error_count = validation.error_count
    job.result = validation.result
    job.status = ImportJob.Status.FAILED if validation.error_count else ImportJob.Status.VALIDATED
    job.save(update_fields=["row_count", "error_count", "result", "status", "updated_at"])

    _create_audit_log(
        actor=actor,
        action="import.vehicle.validated",
        entity_type="import_job",
        entity_id=job.id,
        before={},
        after={"status": job.status, "row_count": job.row_count, "error_count": job.error_count},
        request_meta=request_meta,
    )
    return job


def validate_vehicle_workbook(uploaded_file, mapping: dict[str, Any] | None = None) -> ImportValidationResult:
    """Validate a vehicle workbook and return normalized rows ready for commit.

    When ``mapping`` (a dict of ``{column_key: source_column_index}``) is given,
    it overrides the automatic header detection so the user can assign columns
    interactively. Otherwise headers are matched automatically via aliases.
    """
    filename = uploaded_file.name or ""
    if Path(filename).suffix.lower() not in SUPPORTED_EXTENSIONS:
        return _file_error(_("Only .xlsx or .xlsm files are supported."))

    try:
        workbook = load_workbook(BytesIO(uploaded_file.read()), read_only=True, data_only=True)
    except (InvalidFileException, OSError, ValueError):
        return _file_error(_("The uploaded Excel file could not be read."))
    finally:
        try:
            uploaded_file.seek(0)
        except (AttributeError, OSError):
            pass

    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration:
        workbook.close()
        return _file_error(_("The workbook must contain a header row."))

    # Peek at the first data row so the mapping UI can show example values, then
    # put it back so it is still validated below.
    first_data_row = next(rows, None)
    if first_data_row is not None:
        rows = itertools.chain([first_data_row], rows)
    source_columns = _source_columns(raw_headers, first_data_row or ())

    auto_header_map = _build_header_map(raw_headers)
    header_map = _header_map_from_mapping(mapping, raw_headers) if mapping is not None else auto_header_map

    def _with_mapping_meta(result: dict[str, Any]) -> dict[str, Any]:
        result["source_columns"] = source_columns
        result["suggested_mapping"] = auto_header_map
        result["mapping"] = header_map
        return result

    header_errors = _validate_headers(header_map)
    if header_errors:
        workbook.close()
        result = _with_mapping_meta(_empty_result())
        result["errors"] = header_errors
        return ImportValidationResult(row_count=0, error_count=len(header_errors), result=result)

    category_by_name = {
        _normalize_lookup(category.name): category
        for category in VehicleCategory.objects.filter(is_active=True).only("id", "name")
    }
    vehicle_by_internal = {
        vehicle.internal_number: vehicle
        for vehicle in Vehicle.objects.only(
            "id",
            "internal_number",
            "serial_number",
            "license_plate",
            "current_odometer_km",
            "current_operating_hours",
        )
    }

    parsed_rows: list[dict[str, Any]] = []
    seen_internal_numbers: dict[str, int] = {}
    seen_serial_numbers: dict[str, int] = {}
    seen_license_plates: dict[str, int] = {}
    row_count = 0
    error_count = 0

    for row_number, raw_row in enumerate(rows, start=2):
        values = _row_values(raw_row, header_map)
        if _is_empty_row(values):
            continue

        row_count += 1
        normalized_values, errors = _normalize_row(values, category_by_name)
        action = "update" if normalized_values.get("internal_number") in vehicle_by_internal else "create"

        errors.extend(
            _validate_row_uniqueness(
                normalized_values,
                row_number,
                vehicle_by_internal,
                seen_internal_numbers,
                seen_serial_numbers,
                seen_license_plates,
            )
        )
        errors.extend(_validate_reading_decreases(normalized_values, vehicle_by_internal))

        parsed_rows.append(
            {
                "row_number": row_number,
                "action": action,
                "values": _serializable_values(normalized_values),
                "errors": errors,
            }
        )
        error_count += len(errors)

    workbook.close()
    result = _with_mapping_meta(_empty_result())
    result["rows"] = parsed_rows
    result["summary"] = {
        "row_count": row_count,
        "error_count": error_count,
        "create_count": sum(1 for row in parsed_rows if row["action"] == "create" and not row["errors"]),
        "update_count": sum(1 for row in parsed_rows if row["action"] == "update" and not row["errors"]),
    }
    return ImportValidationResult(row_count=row_count, error_count=error_count, result=result)


@transaction.atomic
def revalidate_vehicle_import_job(
    *, job: ImportJob, mapping: dict[str, Any] | None, actor, request_meta: dict[str, str]
) -> ImportJob:
    """Re-run validation for an existing job with a user-supplied column mapping."""
    job = ImportJob.objects.select_for_update().get(pk=job.pk)
    if job.import_type != ImportJob.ImportType.VEHICLES:
        raise ValueError(_("Only vehicle import jobs can be remapped by this endpoint."))
    if job.status == ImportJob.Status.COMMITTED:
        raise ValueError(_("This import job has already been committed."))

    storage_key = job.source_media.storage_key
    if not default_storage.exists(storage_key):
        raise ValueError(_("The original import file is no longer available."))

    with default_storage.open(storage_key, "rb") as source_file:
        validation = validate_vehicle_workbook(source_file, mapping=mapping)

    job.row_count = validation.row_count
    job.error_count = validation.error_count
    job.result = validation.result
    job.status = ImportJob.Status.FAILED if validation.error_count else ImportJob.Status.VALIDATED
    job.save(update_fields=["row_count", "error_count", "result", "status", "updated_at"])

    _create_audit_log(
        actor=actor,
        action="import.vehicle.remapped",
        entity_type="import_job",
        entity_id=job.id,
        before={},
        after={"status": job.status, "row_count": job.row_count, "error_count": job.error_count},
        request_meta=request_meta,
    )
    return job


@transaction.atomic
def commit_vehicle_import_job(*, job: ImportJob, actor, request_meta: dict[str, str]) -> ImportJob:
    """Create/update vehicles for a previously validated ImportJob."""
    job = ImportJob.objects.select_for_update().get(pk=job.pk)
    if job.import_type != ImportJob.ImportType.VEHICLES:
        raise ValueError(_("Only vehicle import jobs can be committed by this endpoint."))
    if job.status == ImportJob.Status.COMMITTED:
        raise ValueError(_("This import job has already been committed."))
    if job.status != ImportJob.Status.VALIDATED or job.error_count:
        raise ValueError(_("Import job must be successfully validated before commit."))

    created_count = 0
    updated_count = 0
    committed_rows: list[dict[str, Any]] = []

    fallback_category: VehicleCategory | None = None

    for row in job.result.get("rows", []):
        values = row["values"]
        category_id = values.get("category_id")
        if category_id:
            category = VehicleCategory.objects.get(pk=category_id)
        else:
            if fallback_category is None:
                fallback_category, _created = VehicleCategory.objects.get_or_create(
                    name=FALLBACK_CATEGORY_NAME,
                    defaults={"is_active": True},
                )
            category = fallback_category
        vehicle = Vehicle.objects.filter(internal_number=values["internal_number"]).first()
        before = _vehicle_snapshot(vehicle) if vehicle else {}

        vehicle_values = {
            "category": category,
            "manufacturer": values["manufacturer"],
            "model": values["model"],
            "serial_number": values.get("serial_number", ""),
            "license_plate": values.get("license_plate", ""),
            "current_odometer_km": values.get("current_odometer_km"),
            "current_operating_hours": _decimal_or_none(values.get("current_operating_hours")),
            "current_location": values.get("current_location", ""),
            "notes": values.get("notes", ""),
        }
        try:
            if vehicle:
                for field, value in vehicle_values.items():
                    setattr(vehicle, field, value)
                vehicle.save()
                updated_count += 1
                action = "update"
            else:
                vehicle = Vehicle.objects.create(
                    internal_number=values["internal_number"],
                    status=VehicleStatus.AVAILABLE,
                    **vehicle_values,
                )
                created_count += 1
                action = "create"
        except (DjangoValidationError, IntegrityError) as exc:
            raise ValueError(
                _("Row %(row)s could not be imported: %(error)s")
                % {"row": row["row_number"], "error": _humanize_save_error(exc)}
            ) from exc

        committed_rows.append({"row_number": row["row_number"], "action": action, "vehicle_id": str(vehicle.id)})
        _create_audit_log(
            actor=actor,
            action=f"import.vehicle.{action}d",
            entity_type="vehicle",
            entity_id=vehicle.id,
            before=before,
            after=_vehicle_snapshot(vehicle),
            request_meta=request_meta,
        )

    result = dict(job.result)
    result["commit"] = {
        "created_count": created_count,
        "updated_count": updated_count,
        "committed_rows": committed_rows,
    }
    job.status = ImportJob.Status.COMMITTED
    job.committed_at = timezone.now()
    job.result = result
    job.save(update_fields=["status", "committed_at", "result", "updated_at"])

    _create_audit_log(
        actor=actor,
        action="import.vehicle.committed",
        entity_type="import_job",
        entity_id=job.id,
        before={"status": ImportJob.Status.VALIDATED},
        after={"status": job.status, "created_count": created_count, "updated_count": updated_count},
        request_meta=request_meta,
    )
    return job


def _empty_result() -> dict[str, Any]:
    return {
        "columns": EXPECTED_COLUMNS,
        "required_columns": REQUIRED_COLUMNS,
        "source_columns": [],
        "suggested_mapping": {},
        "mapping": {},
        "rows": [],
        "errors": [],
        "summary": {"row_count": 0, "error_count": 0, "create_count": 0, "update_count": 0},
    }


def _source_columns(raw_headers: tuple[Any, ...], sample_row: tuple[Any, ...]) -> list[dict[str, Any]]:
    """Describe the spreadsheet's own columns for the interactive mapping UI."""
    columns: list[dict[str, Any]] = []
    for index, label in enumerate(raw_headers):
        if label is None or str(label).strip() == "":
            continue
        sample = ""
        if index < len(sample_row) and sample_row[index] is not None:
            sample = _clean_text(sample_row[index])
        columns.append({"index": index, "label": str(label).strip(), "sample": sample})
    return columns


def _header_map_from_mapping(mapping: dict[str, Any] | None, raw_headers: tuple[Any, ...]) -> dict[str, int]:
    """Build a column->index map from an explicit user-supplied mapping."""
    header_map: dict[str, int] = {}
    if not isinstance(mapping, dict):
        return header_map
    header_len = len(raw_headers)
    used_indices: set[int] = set()
    for column in EXPECTED_COLUMNS:
        raw_index = mapping.get(column)
        if raw_index in (None, ""):
            continue
        try:
            index = int(raw_index)
        except (TypeError, ValueError):
            continue
        if 0 <= index < header_len and index not in used_indices:
            header_map[column] = index
            used_indices.add(index)
    return header_map


def _file_error(message: str) -> ImportValidationResult:
    result = _empty_result()
    result["errors"] = [{"code": "invalid_file", "message": str(message)}]
    result["summary"]["error_count"] = 1
    return ImportValidationResult(row_count=0, error_count=1, result=result)


def _normalize_header(value: Any) -> str:
    """Normalize a raw header cell for tolerant matching.

    Lower-cases, trims, and collapses spaces/hyphens/dots/slashes into single
    underscores so headers like "Internal Number", "interne-nummer" or
    "KFZ Kennzeichen" all resolve to a canonical column key.
    """
    text = str(value).strip().lower()
    for separator in (" ", "-", "/", ".", "\\"):
        text = text.replace(separator, "_")
    while "__" in text:
        text = text.replace("__", "_")
    return text.strip("_")


def _build_header_map(raw_headers: tuple[Any, ...]) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for index, value in enumerate(raw_headers):
        if value is None:
            continue
        column = _ALIAS_TO_COLUMN.get(_normalize_header(value))
        if column is not None and column not in header_map:
            header_map[column] = index
    return header_map


def _validate_headers(header_map: dict[str, int]) -> list[dict[str, str]]:
    errors = []
    for column in REQUIRED_COLUMNS:
        if column not in header_map:
            errors.append(
                {
                    "code": "missing_required_column",
                    "field": column,
                    "message": str(_("Missing required column: %(column)s.") % {"column": column}),
                }
            )
    return errors


def _row_values(raw_row: tuple[Any, ...], header_map: dict[str, int]) -> dict[str, Any]:
    values = {}
    for column in EXPECTED_COLUMNS:
        index = header_map.get(column)
        values[column] = raw_row[index] if index is not None and index < len(raw_row) else None
    return values


def _is_empty_row(values: dict[str, Any]) -> bool:
    return all(_clean_text(value) == "" for value in values.values())


def _normalize_row(
    values: dict[str, Any],
    category_by_name: dict[str, VehicleCategory],
) -> tuple[dict[str, Any], list[dict[str, str]]]:
    errors: list[dict[str, str]] = []
    normalized: dict[str, Any] = {
        "internal_number": _clean_text(values["internal_number"]),
        "category": _clean_text(values["category"]),
        "manufacturer": _clean_text(values["manufacturer"]),
        "model": _clean_text(values["model"]),
        "serial_number": _clean_text(values["serial_number"]),
        "license_plate": _clean_text(values["license_plate"]),
        "current_location": _clean_text(values["current_location"]),
        "supplier": _clean_text(values["supplier"]),
        "notes": _clean_text(values["notes"]),
    }

    for field in REQUIRED_COLUMNS:
        if not normalized[field]:
            errors.append(_field_error(field, "required", _("%(field)s is required.") % {"field": field}))

    for field, limit in _MAX_FIELD_LENGTHS.items():
        value = normalized.get(field, "")
        if value and len(value) > limit:
            errors.append(
                _field_error(
                    field,
                    "too_long",
                    _("%(field)s must be at most %(limit)d characters.") % {"field": field, "limit": limit},
                )
            )

    category = category_by_name.get(_normalize_lookup(normalized["category"]))
    if category:
        normalized["category_id"] = str(category.id)
    else:
        # Unknown or blank category: route the vehicle into the catch-all
        # category at commit time instead of failing the row.
        normalized["category_fallback"] = True

    normalized["current_odometer_km"] = _non_negative_int(
        values["current_odometer_km"],
        "current_odometer_km",
        errors,
    )
    normalized["current_operating_hours"] = _non_negative_decimal(
        values["current_operating_hours"], "current_operating_hours", errors
    )
    return normalized, errors


def _validate_row_uniqueness(
    values: dict[str, Any],
    row_number: int,
    vehicle_by_internal: dict[str, Vehicle],
    seen_internal_numbers: dict[str, int],
    seen_serial_numbers: dict[str, int],
    seen_license_plates: dict[str, int],
) -> list[dict[str, str]]:
    errors: list[dict[str, str]] = []
    internal_number = values.get("internal_number", "")
    if internal_number:
        if internal_number in seen_internal_numbers:
            errors.append(
                _field_error(
                    "internal_number",
                    "duplicate_in_file",
                    _("Duplicate internal_number in import file. First seen on row %(row)s.")
                    % {"row": seen_internal_numbers[internal_number]},
                )
            )
        else:
            seen_internal_numbers[internal_number] = row_number

    vehicle = vehicle_by_internal.get(internal_number)
    errors.extend(
        _validate_unique_optional_value(
            field="serial_number",
            value=values.get("serial_number", ""),
            current_vehicle=vehicle,
            seen_values=seen_serial_numbers,
            row_number=row_number,
            conflict_message=_("Serial number already belongs to another vehicle."),
            duplicate_message=_("Duplicate serial_number in import file. First seen on row %(row)s."),
        )
    )
    errors.extend(
        _validate_unique_optional_value(
            field="license_plate",
            value=values.get("license_plate", ""),
            current_vehicle=vehicle,
            seen_values=seen_license_plates,
            row_number=row_number,
            conflict_message=_("License plate already belongs to another vehicle."),
            duplicate_message=_("Duplicate license_plate in import file. First seen on row %(row)s."),
        )
    )
    return errors


def _validate_unique_optional_value(
    *,
    field: str,
    value: str,
    current_vehicle: Vehicle | None,
    seen_values: dict[str, int],
    row_number: int,
    conflict_message: str,
    duplicate_message: str,
) -> list[dict[str, str]]:
    if not value:
        return []

    errors: list[dict[str, str]] = []
    if value in seen_values:
        errors.append(
            _field_error(field, "duplicate_in_file", duplicate_message % {"row": seen_values[value]})
        )
    else:
        seen_values[value] = row_number

    conflict = Vehicle.objects.filter(**{field: value}).only("id").first()
    if conflict and (current_vehicle is None or conflict.id != current_vehicle.id):
        errors.append(_field_error(field, "unique_conflict", conflict_message))
    return errors


def _validate_reading_decreases(
    values: dict[str, Any],
    vehicle_by_internal: dict[str, Vehicle],
) -> list[dict[str, str]]:
    vehicle = vehicle_by_internal.get(values.get("internal_number", ""))
    if not vehicle:
        return []

    errors: list[dict[str, str]] = []
    odometer = values.get("current_odometer_km")
    if vehicle.current_odometer_km is not None and odometer is not None and odometer < vehicle.current_odometer_km:
        errors.append(
            _field_error("current_odometer_km", "decreasing_reading", _("Odometer value must not decrease."))
        )

    hours = _decimal_or_none(values.get("current_operating_hours"))
    if vehicle.current_operating_hours is not None and hours is not None and hours < vehicle.current_operating_hours:
        errors.append(
            _field_error("current_operating_hours", "decreasing_reading", _("Operating hours must not decrease."))
        )
    return errors


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return format(value, "f").rstrip("0").rstrip(".")
    return str(value).strip()


def _normalize_lookup(value: str) -> str:
    return value.strip().casefold()


def _non_negative_int(value: Any, field: str, errors: list[dict[str, str]]) -> int | None:
    if value is None or _clean_text(value) == "":
        return None
    try:
        if isinstance(value, bool):
            raise ValueError
        decimal_value = Decimal(str(value).strip())
        if decimal_value != decimal_value.to_integral_value() or decimal_value < 0:
            raise ValueError
        return int(decimal_value)
    except (InvalidOperation, ValueError):
        errors.append(
            _field_error(field, "invalid_integer", _("%(field)s must be a non-negative integer.") % {"field": field})
        )
        return None


def _non_negative_decimal(value: Any, field: str, errors: list[dict[str, str]]) -> str | None:
    if value is None or _clean_text(value) == "":
        return None
    try:
        if isinstance(value, bool):
            raise InvalidOperation
        decimal_value = Decimal(str(value).strip()).quantize(Decimal("0.1"))
        if decimal_value < 0:
            raise InvalidOperation
        return str(decimal_value)
    except (InvalidOperation, ValueError):
        errors.append(
            _field_error(
                field,
                "invalid_decimal",
                _("%(field)s must be a non-negative decimal number.") % {"field": field},
            )
        )
        return None


def _decimal_or_none(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    return Decimal(str(value))


def _serializable_values(values: dict[str, Any]) -> dict[str, Any]:
    return {key: (str(value) if isinstance(value, Decimal) else value) for key, value in values.items()}


def _humanize_save_error(exc: Exception) -> str:
    """Turn a model/database save error into a readable, field-aware message."""
    if isinstance(exc, DjangoValidationError):
        parts: list[str] = []
        message_dict = getattr(exc, "message_dict", None)
        if message_dict:
            for field, messages in message_dict.items():
                joined = " ".join(str(message) for message in messages)
                parts.append(f"{field}: {joined}" if field and field != "__all__" else joined)
        else:
            parts.extend(str(message) for message in getattr(exc, "messages", [str(exc)]))
        return " ".join(part for part in parts if part) or str(exc)
    # IntegrityError and friends: surface a concise, non-internal message.
    return _("A database constraint was violated (e.g. a duplicate serial number or license plate).")


def _field_error(field: str, code: str, message: str) -> dict[str, str]:
    return {"field": field, "code": code, "message": str(message)}


def _vehicle_snapshot(vehicle: Vehicle | None) -> dict[str, Any]:
    if vehicle is None:
        return {}
    return {
        "id": str(vehicle.id),
        "internal_number": vehicle.internal_number,
        "category": str(vehicle.category_id),
        "manufacturer": vehicle.manufacturer,
        "model": vehicle.model,
        "serial_number": vehicle.serial_number,
        "license_plate": vehicle.license_plate,
        "status": vehicle.status,
        "current_odometer_km": vehicle.current_odometer_km,
        "current_operating_hours": (
            str(vehicle.current_operating_hours) if vehicle.current_operating_hours is not None else None
        ),
        "current_location": vehicle.current_location,
        "notes": vehicle.notes,
    }


def _create_audit_log(
    *,
    actor,
    action: str,
    entity_type: str,
    entity_id,
    before: dict[str, Any],
    after: dict[str, Any],
    request_meta: dict[str, str],
) -> None:
    AuditLog.objects.create(
        actor=actor,
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        before=before,
        after=after,
        ip_address=request_meta.get("ip_address") or None,
        user_agent=request_meta.get("user_agent", ""),
    )
